#!/usr/bin/env python3
"""
Demultiplex filtered FASTQ using barcodes from barcode.xlsx and optionally library.xlsx.
"""
import argparse
import os
import sys
import logging
import openpyxl
from Bio import SeqIO

def setup_logger(name, log_dir="./logs", level=logging.INFO):
    os.makedirs(log_dir, exist_ok=True)
    logger = logging.getLogger(name)
    logger.setLevel(level)
    log_file = os.path.join(log_dir, f"{name}.log")
    fh = logging.FileHandler(log_file)
    fh.setLevel(level)
    ch = logging.StreamHandler()
    ch.setLevel(level)
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    fh.setFormatter(formatter)
    ch.setFormatter(formatter)
    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--forward", required=True, help="Filtered forward FASTQ")
    parser.add_argument("--reverse", required=True, help="Filtered reverse FASTQ")
    parser.add_argument("--barcode", required=True, help="barcode.xlsx file")
    parser.add_argument("--library", default=None, help="library.xlsx file (sample name to password mapping, optional)")
    parser.add_argument("--output_dir", required=True, help="Output directory for demultiplexed FASTQ")
    parser.add_argument("--log-level", default="INFO", choices=["DEBUG","INFO","WARNING","ERROR"])
    args = parser.parse_args()

    log_level = getattr(logging, args.log_level.upper())
    logger = setup_logger("demultiplex", level=log_level)

    try:
        # Check input files
        for f in [args.forward, args.reverse, args.barcode]:
            if not os.path.exists(f):
                raise FileNotFoundError(f"Required file not found: {f}")

        # Read library mapping if provided
        name_pass = {}
        if args.library and os.path.exists(args.library):
            logger.info("Reading library mapping...")
            wb = openpyxl.load_workbook(args.library)
            ws = wb.active
            for row in range(2, ws.max_row + 1):
                name = ws.cell(row, 1).value
                pwd = ws.cell(row, 2).value
                if name and pwd:
                    name_pass[name] = pwd
            logger.info(f"Loaded {len(name_pass)} name-password mappings.")
        else:
            logger.info("No library file provided or file not found; skipping name replacement.")

        # Process barcode.xlsx: replace sample names with passwords if mapping exists
        if name_pass:
            logger.info("Processing barcode.xlsx...")
            wb_bc = openpyxl.load_workbook(args.barcode)
            ws_bc = wb_bc.active
            for row in range(2, ws_bc.max_row + 1):
                for col in range(1, 4):
                    val = ws_bc.cell(row, col).value
                    if val in name_pass:
                        ws_bc.cell(row, col).value = name_pass[val]
            wb_bc.save(args.barcode)
            logger.info("Updated barcode.xlsx with passwords.")
        else:
            logger.info("No name replacement performed.")

        # Generate barcode mapping from barcode.xlsx
        wb_bc = openpyxl.load_workbook(args.barcode)
        ws_bc = wb_bc.active
        barcode_map = {}
        for row in range(2, ws_bc.max_row + 1):
            sample = ws_bc.cell(row, 1).value
            idx1 = ws_bc.cell(row, 2).value
            idx2 = ws_bc.cell(row, 3).value
            if sample and idx1 and idx2:
                barcode_map[sample] = (idx1, idx2)
        logger.info(f"Loaded {len(barcode_map)} barcode pairs.")

        # Create output directory
        os.makedirs(args.output_dir, exist_ok=True)

        # Demultiplex
        logger.info("Demultiplexing...")
        fwd_iter = SeqIO.parse(args.forward, 'fastq')
        rev_iter = SeqIO.parse(args.reverse, 'fastq')

        processed = 0
        matched = 0
        unmatched = 0

        for f_rec, r_rec in zip(fwd_iter, rev_iter):
            processed += 1
            f_idx = str(f_rec.seq[:10])
            r_idx = str(r_rec.seq[:10])
            found = False
            for sample, (exp1, exp2) in barcode_map.items():
                if f_idx == exp1 and r_idx == exp2:
                    f_out = os.path.join(args.output_dir, f"{sample}_1.fastq")
                    r_out = os.path.join(args.output_dir, f"{sample}_2.fastq")
                    with open(f_out, 'a') as fh:
                        SeqIO.write(f_rec, fh, 'fastq')
                    with open(r_out, 'a') as rh:
                        SeqIO.write(r_rec, rh, 'fastq')
                    matched += 1
                    found = True
                    break
            if not found:
                unmatched += 1
                if processed % 10000 == 0:
                    logger.debug(f"Unmatched record at {processed}")

        logger.info(f"Processed {processed} records. Matched: {matched}, Unmatched: {unmatched}.")
        logger.info("Demultiplexing completed.")

    except Exception as e:
        logger.error(f"Demultiplexing failed: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()